"""
Excel file processing service
"""
import openpyxl
from datetime import datetime
from models import Task, db


class ExcelProcessor:
    """Process Excel files for task import"""

    def __init__(self):
        self.errors = []
        self.stats = {
            'total_rows': 0,
            'valid_rows': 0,
            'invalid_rows': 0,
            'new_records': 0,
            'duplicates_skipped': 0
        }

    def parse_excel(self, file_path):
        """
        Parse Excel file and extract tasks
        Returns list of task dictionaries
        """
        tasks = []
        self.errors = []

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Get total rows
            self.stats['total_rows'] = ws.max_row - 2  # Exclude header rows

            # Read data starting from row 3 (assuming row 1-2 are headers)
            current_stt = None
            current_department = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not any(row):  # Skip completely empty rows
                    continue

                try:
                    stt = row[0] if row[0] else current_stt
                    department = row[1] if row[1] else current_department
                    content = row[2]
                    warning_date = row[3]

                    # Update current values for merged cells
                    if row[0]:
                        current_stt = row[0]
                    if row[1]:
                        current_department = row[1]

                    # Validate required fields
                    if not content:
                        self.errors.append(f"Row {row_idx}: Missing content")
                        self.stats['invalid_rows'] += 1
                        continue

                    if not department:
                        self.errors.append(f"Row {row_idx}: Missing department")
                        self.stats['invalid_rows'] += 1
                        continue

                    # Parse date
                    parsed_date = None
                    if warning_date:
                        if isinstance(warning_date, datetime):
                            parsed_date = warning_date.date()
                        elif isinstance(warning_date, str):
                            try:
                                parsed_date = datetime.strptime(warning_date, '%Y-%m-%d').date()
                            except ValueError:
                                self.errors.append(f"Row {row_idx}: Invalid date format")

                    # Create task dictionary
                    task_dict = {
                        'stt': int(stt) if stt and str(stt).isdigit() else None,
                        'department': str(department).strip(),
                        'content': str(content).strip(),
                        'warning_date': parsed_date
                    }

                    tasks.append(task_dict)
                    self.stats['valid_rows'] += 1

                except Exception as e:
                    self.errors.append(f"Row {row_idx}: {str(e)}")
                    self.stats['invalid_rows'] += 1
                    continue

            wb.close()

        except Exception as e:
            self.errors.append(f"Error reading Excel file: {str(e)}")
            raise

        return tasks

    def merge_tasks(self, new_tasks):
        """
        Merge new tasks with existing tasks in database
        Skip duplicates based on: department + content + warning_date
        Returns statistics about the merge operation
        """
        # Get all existing tasks
        existing_tasks = Task.query.all()

        # Create set of unique keys for existing tasks
        existing_keys = set()
        for task in existing_tasks:
            key = self._create_task_key(task.department, task.content, task.warning_date)
            existing_keys.add(key)

        # Filter out duplicates from new tasks
        tasks_to_add = []
        for new_task in new_tasks:
            key = self._create_task_key(
                new_task['department'],
                new_task['content'],
                new_task['warning_date']
            )

            if key in existing_keys:
                self.stats['duplicates_skipped'] += 1
            else:
                tasks_to_add.append(new_task)
                existing_keys.add(key)  # Add to set to avoid duplicates within new tasks

        # Add new tasks to database
        for task_dict in tasks_to_add:
            task = Task(**task_dict)
            db.session.add(task)
            self.stats['new_records'] += 1

        db.session.commit()

        return self.stats

    def _create_task_key(self, department, content, warning_date):
        """Create unique key for task deduplication"""
        date_str = warning_date.isoformat() if warning_date else 'no_date'
        return f"{department}|{content}|{date_str}"

    def get_stats(self):
        """Get import statistics"""
        return self.stats

    def get_errors(self):
        """Get list of errors encountered during processing"""
        return self.errors
